"""Leitura da aba ``Tabela de Transferência`` e extração dos dados.

Responsável por:

* montar o mapa ``chave -> valor cru`` (aplicando a política de conflito);
* detectar e reportar conflitos de chaves duplicadas com valores divergentes;
* extrair, **por chave estável** (não por número de linha fixo), os dados dos
  dois gráficos, validando de quebra os marcadores da coluna F.

Lê com ``data_only=True`` para obter os valores calculados (não as fórmulas).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from numbers import Real
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import Config

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
# Estruturas de dados
# --------------------------------------------------------------------------- #
@dataclass
class KeyOccurrence:
    row: int
    raw_value: object


@dataclass
class Conflict:
    """Chave duplicada com valores divergentes."""

    key: str
    occurrences: list[KeyOccurrence]
    chosen_row: int
    chosen_value: object

    def describe(self) -> str:
        parts = ", ".join(f"linha {o.row}={o.raw_value!r}" for o in self.occurrences)
        return (
            f"{self.key}: valores divergentes [{parts}] -> escolhido "
            f"linha {self.chosen_row}={self.chosen_value!r}"
        )


@dataclass
class PieData:
    """Dados do Gráfico 1 (pizza)."""

    labels: list[str]
    values: list[float]
    dropped_zero: list[str] = field(default_factory=list)
    total_key: str = ""
    total_value: float | None = None


@dataclass
class BarData:
    """Dados do Gráfico 2 (barras mensais)."""

    months: list[str]
    values: list[float]
    dropped: list[str] = field(default_factory=list)  # descrições de descartes


@dataclass
class StackedBarData:
    """Dados do Gráfico 3 (barras empilhadas: Ponta + Fora Ponta)."""

    months: list[str]
    ponta: list[float]
    fora_ponta: list[float]
    dropped: list[str] = field(default_factory=list)


@dataclass
class DemandData:
    """Dados do Gráfico 4 (demanda Ponta/Fora Ponta + linhas de contrato)."""

    months: list[str]
    ponta: list[float]
    fora_ponta: list[float]
    contratada: float | None = None
    tolerancia: float | None = None
    tolerancia_np: float | None = None
    tolerancia_fp: float | None = None
    dropped: list[str] = field(default_factory=list)


@dataclass
class ExpenseData:
    """Dados do Gráfico 5 (despesa mensal com energia)."""

    months: list[str]
    values: list[float]
    dropped: list[str] = field(default_factory=list)


@dataclass
class LoadResult:
    resolved: dict[str, object]
    occurrences: dict[str, list[KeyOccurrence]]
    conflicts: list[Conflict]
    planilha_keys: set[str]
    pie: PieData
    bar: BarData
    stacked: StackedBarData
    demand: DemandData
    expense: ExpenseData
    marker_warnings: list[str]

    @property
    def duplicate_keys(self) -> dict[str, list[KeyOccurrence]]:
        return {k: v for k, v in self.occurrences.items() if len(v) > 1}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _to_float(value: object) -> float | None:
    """Coerção tolerante para float. Retorna None se não for numérico."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Real):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(".", "").replace(",", ".")  # tenta pt-BR
        try:
            return float(s)
        except ValueError:
            return None
    return None


# --------------------------------------------------------------------------- #
# Parsing das substituições
# --------------------------------------------------------------------------- #
def parse_substitutions(
    ws: Worksheet, config: Config
) -> tuple[dict[str, object], dict[str, list[KeyOccurrence]], list[Conflict]]:
    """Percorre a aba e monta ``resolved``, ``occurrences`` e ``conflicts``."""
    key_re = re.compile(config.key_pattern)
    occurrences: dict[str, list[KeyOccurrence]] = {}

    for row in range(config.header_row + 1, ws.max_row + 1):
        key_cell = ws[f"{config.col_key}{row}"].value
        if not isinstance(key_cell, str):
            continue
        key = key_cell.strip()
        if not key_re.match(key):
            continue
        value = ws[f"{config.col_value}{row}"].value
        occurrences.setdefault(key, []).append(KeyOccurrence(row, value))

    resolved: dict[str, object] = {}
    conflicts: list[Conflict] = []

    for key, occs in occurrences.items():
        non_empty = [o for o in occs if not _is_empty(o.raw_value)]
        pool = non_empty or occs  # se tudo vazio, usa as ocorrências cruas

        if config.conflict_policy == "first-wins":
            chosen = pool[0]
        else:  # last-wins (default)
            chosen = pool[-1]
        resolved[key] = chosen.raw_value

        # Conflito = duplicada COM valores divergentes (ignora duplicatas iguais).
        distinct = {repr(o.raw_value) for o in non_empty}
        if len(occs) > 1 and len(distinct) > 1:
            conflict = Conflict(key, occs, chosen.row, chosen.raw_value)
            conflicts.append(conflict)
            logger.warning("Conflito de chave duplicada -> %s", conflict.describe())
        elif len(occs) > 1:
            logger.debug("Chave duplicada (mesmo valor), sem conflito: %s", key)

    logger.info(
        "Substituições: %d chaves distintas, %d duplicadas, %d conflitos reais.",
        len(resolved),
        sum(1 for v in occurrences.values() if len(v) > 1),
        len(conflicts),
    )
    return resolved, occurrences, conflicts


# --------------------------------------------------------------------------- #
# Extração dos gráficos
# --------------------------------------------------------------------------- #
def extract_pie(resolved: dict[str, object], config: Config) -> PieData:
    """Gráfico 1: fatias por uso final (exclui o total, omite fatias zero)."""
    labels: list[str] = []
    values: list[float] = []
    dropped: list[str] = []

    for slice_ in config.graph1_slices:
        raw = resolved.get(slice_.key)
        val = _to_float(raw)
        if val is None:
            logger.warning(
                "Gráfico 1: valor não-numérico/ausente para %s (%r); usando 0.",
                slice_.key,
                raw,
            )
            val = 0.0
        if config.graph1_drop_zero_slices and val <= 0:
            dropped.append(slice_.label)
            logger.info("Gráfico 1: fatia '%s' omitida (valor %s).", slice_.label, val)
            continue
        labels.append(slice_.label)
        values.append(val)

    total_value = _to_float(resolved.get(config.graph1_total_key))
    logger.info(
        "Gráfico 1: %d fatias plotadas %s; total (%s) EXCLUÍDO da pizza.",
        len(values),
        dict(zip(labels, values)),
        config.graph1_total_key,
    )
    return PieData(
        labels=labels,
        values=values,
        dropped_zero=dropped,
        total_key=config.graph1_total_key,
        total_value=total_value,
    )


def extract_bar(resolved: dict[str, object], config: Config) -> BarData:
    """Gráfico 2: consumo mensal, descartando meses 'Sem Histórico Disponível'."""
    months: list[str] = []
    values: list[float] = []
    dropped: list[str] = []

    no_hist = config.graph2_no_history_label.strip().lower()
    pairs = zip(config.graph2_month_keys, config.graph2_value_keys)
    for idx, (mkey, vkey) in enumerate(pairs, start=1):
        month_raw = resolved.get(mkey)
        month = "" if month_raw is None else str(month_raw).strip()
        val = _to_float(resolved.get(vkey)) or 0.0

        if month.lower() == no_hist:
            dropped.append(f"mês {idx} ({mkey}): sem histórico")
            continue
        if config.graph2_drop_zero_values and val == 0:
            dropped.append(f"mês {idx} ({month}): consumo zero")
            continue
        months.append(month)
        values.append(val)

    if dropped:
        logger.info("Gráfico 2: %d mês(es) descartado(s): %s", len(dropped), dropped)
    logger.info("Gráfico 2: %d mês(es) plotado(s): %s", len(months), dict(zip(months, values)))
    return BarData(months=months, values=values, dropped=dropped)


def _extract_month_series(
    resolved: dict[str, object], config: Config, value_key_lists: list[list[str]]
) -> tuple[list[str], list[list[float]], list[str]]:
    """Extrai meses + N séries alinhadas posicionalmente (1..12).

    Usa as chaves de mês do Gráfico 2 (``<<mesUm>>`` ..) — compartilhadas pelos
    gráficos 2/3/4/5. Descarta meses rotulados como 'Sem Histórico Disponível'.
    Cada série em ``value_key_lists`` é uma lista de 12 chaves posicionais.
    Valores ausentes/não-numéricos viram 0.0.
    """
    no_hist = config.graph2_no_history_label.strip().lower()
    months: list[str] = []
    series: list[list[float]] = [[] for _ in value_key_lists]
    dropped: list[str] = []
    for idx, mkey in enumerate(config.graph2_month_keys, start=1):
        month_raw = resolved.get(mkey)
        month = "" if month_raw is None else str(month_raw).strip()
        if month.lower() == no_hist:
            dropped.append(f"mês {idx} ({mkey}): sem histórico")
            continue
        months.append(month)
        for s_idx, keys in enumerate(value_key_lists):
            series[s_idx].append(_to_float(resolved.get(keys[idx - 1])) or 0.0)
    return months, series, dropped


def extract_stacked(resolved: dict[str, object], config: Config) -> StackedBarData:
    """Gráfico 3: consumo Ponta vs. Fora Ponta por mês (barras empilhadas)."""
    months, (ponta, fora_ponta), dropped = _extract_month_series(
        resolved, config, [config.graph3_ponta_keys, config.graph3_fora_ponta_keys]
    )
    logger.info("Gráfico 3: %d mês(es); soma Ponta=%s, Fora Ponta=%s",
                len(months), sum(ponta), sum(fora_ponta))
    return StackedBarData(months=months, ponta=ponta, fora_ponta=fora_ponta, dropped=dropped)


def extract_demand(resolved: dict[str, object], config: Config) -> DemandData:
    """Gráfico 4: demanda Ponta/Fora Ponta por mês + linhas de contrato."""
    months, (ponta, fora_ponta), dropped = _extract_month_series(
        resolved, config, [config.graph4_ponta_keys, config.graph4_fora_ponta_keys]
    )
    contratada = _to_float(resolved.get(config.graph4_contratada_key))
    tolerancia = _to_float(resolved.get(config.graph4_tolerancia_key))
    tolerancia_np = _to_float(resolved.get(config.graph4_tolerancia_np_key))
    tolerancia_fp = _to_float(resolved.get(config.graph4_tolerancia_fp_key))
    logger.info(
        "Gráfico 4: %d mês(es); contratada=%s, tol=%s, tolNP=%s, tolFP=%s",
        len(months), contratada, tolerancia, tolerancia_np, tolerancia_fp,
    )
    return DemandData(
        months=months, ponta=ponta, fora_ponta=fora_ponta,
        contratada=contratada, tolerancia=tolerancia,
        tolerancia_np=tolerancia_np, tolerancia_fp=tolerancia_fp, dropped=dropped,
    )


def extract_expense(resolved: dict[str, object], config: Config) -> ExpenseData:
    """Gráfico 5: despesa mensal com energia elétrica (barras)."""
    months, (values,), dropped = _extract_month_series(
        resolved, config, [config.graph5_value_keys]
    )
    logger.info("Gráfico 5: %d mês(es); soma despesa=%s", len(months), sum(values))
    return ExpenseData(months=months, values=values, dropped=dropped)


def validate_markers(ws: Worksheet, config: Config) -> list[str]:
    """Confere os marcadores da coluna F (F47='Grafico 1', F101='Grafico 2').

    A extração é por chave (robusta), então divergências aqui geram apenas
    aviso — não quebram o processamento.
    """
    warnings: list[str] = []
    expected_markers = (
        (config.graph1_marker, config.graph1_marker_expected_row),
        (config.graph2_marker, config.graph2_marker_expected_row),
        (config.graph3_marker, config.graph3_marker_expected_row),
        (config.graph4_marker, config.graph4_marker_expected_row),
        (config.graph5_marker, config.graph5_marker_expected_row),
    )
    found: dict[str, list[int]] = {m: [] for m, _ in expected_markers}
    for row in range(1, ws.max_row + 1):
        val = ws[f"{config.col_marker}{row}"].value
        if isinstance(val, str):
            v = val.strip()
            if v in found:
                found[v].append(row)

    for marker, expected in expected_markers:
        rows = found[marker]
        if not rows:
            msg = f"Marcador '{marker}' não encontrado na coluna {config.col_marker}."
            warnings.append(msg)
            logger.warning(msg)
        elif expected not in rows:
            msg = (
                f"Marcador '{marker}' esperado na linha {expected}, "
                f"encontrado em {rows}."
            )
            warnings.append(msg)
            logger.warning(msg)
    return warnings


# --------------------------------------------------------------------------- #
# Orquestração
# --------------------------------------------------------------------------- #
def load(xlsx_path: str | Path, config: Config | None = None) -> LoadResult:
    """Carrega a planilha e devolve tudo que o pipeline precisa."""
    config = config or Config()
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {xlsx_path}")

    logger.info("Lendo planilha: %s", xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)
    if config.sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Aba '{config.sheet_name}' não encontrada. Abas: {wb.sheetnames}"
        )
    ws = wb[config.sheet_name]

    resolved, occurrences, conflicts = parse_substitutions(ws, config)
    marker_warnings = validate_markers(ws, config)
    pie = extract_pie(resolved, config)
    bar = extract_bar(resolved, config)
    stacked = extract_stacked(resolved, config)
    demand = extract_demand(resolved, config)
    expense = extract_expense(resolved, config)

    return LoadResult(
        resolved=resolved,
        occurrences=occurrences,
        conflicts=conflicts,
        planilha_keys=set(resolved.keys()),
        pie=pie,
        bar=bar,
        stacked=stacked,
        demand=demand,
        expense=expense,
        marker_warnings=marker_warnings,
    )
