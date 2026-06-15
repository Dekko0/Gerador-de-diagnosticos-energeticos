# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook

from app import excel_loader
from app.config import Config


def test_loads_expected_key_count(load_result):
    # Nº de chaves distintas na coluna D do exemplo preenchido (CMEI Calabar).
    assert len(load_result.resolved) == 239


def test_fixture_is_conflict_free(load_result):
    # A fixture preenchida atual não tem chaves duplicadas divergentes.
    assert load_result.conflicts == []


def _make_xlsx(path: Path, rows: list[tuple[str, object]]) -> None:
    """Cria uma planilha mínima na aba 'Tabela de Transferência' (col D/E)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela de Transferência"
    ws["D1"], ws["E1"] = "chave", "valor"
    for i, (key, val) in enumerate(rows, start=2):
        ws[f"D{i}"], ws[f"E{i}"] = key, val
    wb.save(path)


def test_conflict_resolution_last_and_first_wins(tmp_path):
    # Chave duplicada com valores divergentes -> conflito + política aplicada.
    p = tmp_path / "conf.xlsx"
    _make_xlsx(p, [("<<x>>", 10), ("<<y>>", 5), ("<<x>>", 20)])
    last = excel_loader.load(p, Config())
    assert [c.key for c in last.conflicts] == ["<<x>>"]
    assert last.resolved["<<x>>"] == 20          # last-wins (default)
    first = excel_loader.load(p, Config(conflict_policy="first-wins"))
    assert first.resolved["<<x>>"] == 10         # first-wins
    # valores iguais duplicados NÃO contam como conflito
    p2 = tmp_path / "noconf.xlsx"
    _make_xlsx(p2, [("<<z>>", 7), ("<<z>>", 7)])
    assert excel_loader.load(p2, Config()).conflicts == []


def test_markers_valid_on_model(model_load_result):
    # O modelo novo tem os 5 marcadores nas linhas esperadas
    # (F47=Gráfico 1, F77=Gráfico 3, F102=Gráfico 2, F115=Gráfico 4, F151=Gráfico 5).
    assert model_load_result.marker_warnings == []


def test_model_has_new_graph_keys(model_load_result):
    # As chaves dos novos gráficos (3/4/5) existem no modelo.
    r = model_load_result.resolved
    for k in ("<<ConsumoPontaUm>>", "<<ConsumoForaPontaDoze>>",
              "<<DemandaPontaUm>>", "<<DemandaForaPontaDoze>>",
              "<<despesaEnergiaUm>>", "<<demandaContratadaToleranciaNP>>",
              "<<demandaContratadaToleranciaFP>>"):
        assert k in r


def test_new_graph_series_shapes(model_load_result):
    # Séries posicionais alinhadas com os meses (Ponta/Fora Ponta/Despesa).
    s, d, e = model_load_result.stacked, model_load_result.demand, model_load_result.expense
    assert len(s.ponta) == len(s.fora_ponta) == len(s.months)
    assert len(d.ponta) == len(d.fora_ponta) == len(d.months)
    assert len(e.values) == len(e.months)


def test_pie_excludes_total_and_zero(load_result):
    pie = load_result.pie
    # Motores e Outros = 0 omitidos; total excluído -> restam 3 fatias.
    assert pie.labels == ["Iluminação", "Climatização", "Refrigeração"]
    assert pie.dropped_zero == ["Sistemas Motrizes", "Outros"]
    assert "<<consumoTotal>>" not in pie.labels
    assert abs(pie.values[1] - 68.06293460113977) < 1e-6  # Climatização


def test_bar_has_full_history(load_result):
    bar = load_result.bar
    # Fixture preenchida: 12 meses com histórico, nenhum descartado.
    assert bar.months == ["dez/24", "jan/25", "fev/25", "mar/25", "mai/25",
                          "jul/25", "out/25", "nov/25", "dez/25", "jan/26",
                          "fev/26", "mar/26"]
    assert len(bar.values) == 12
    assert bar.dropped == []
    assert abs(bar.values[0] - 3218.1) < 1e-6


def test_stacked_and_expense_have_data(load_result):
    # Gráficos 3 e 5 têm dados reais na fixture preenchida.
    assert sum(load_result.stacked.ponta) > 0
    assert sum(load_result.stacked.fora_ponta) > 0
    assert sum(load_result.expense.values) > 0


def test_demand_reference_lines(load_result):
    # Gráfico 4: contratada e tolerância única presentes; NP/FP ausentes.
    d = load_result.demand
    assert d.contratada == 110.0
    assert d.tolerancia == 115.5
    assert d.tolerancia_np is None and d.tolerancia_fp is None


def test_planilha_only_keys(load_result):
    # As 11 chaves alimentadoras de gráfico não têm alvo no LaTeX, mas existem.
    for k in ("<<consumoIluminacao>>", "<<propFotovoltaico>>", "<<mesMenorConsumo>>"):
        assert k in load_result.resolved
